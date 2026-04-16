"""
Generador de Layout de Dispersión Omonel  v2.3
Procesa archivos de nómina y cuentas vales para generar dispersión bancaria.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys
import datetime
import threading
import queue
import subprocess
import platform

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

APP_VERSION = "2.3"

# ─────────────────────────── Design Tokens ────────────────────────────

_SYS = platform.system()
_FF  = "Helvetica Neue" if _SYS == "Darwin" else ("Segoe UI" if _SYS == "Windows" else "DejaVu Sans")
_FM  = "SF Mono"        if _SYS == "Darwin" else ("Consolas"  if _SYS == "Windows" else "DejaVu Sans Mono")

FONT_TITLE  = (_FF, 15, "bold")
FONT_LABEL  = (_FF,  9, "bold")
FONT_SMALL  = (_FF,  9)
FONT_MONO   = (_FM,  9)
FONT_BTN    = (_FF, 10, "bold")
FONT_BADGE  = (_FF,  9, "bold")
FONT_NUM    = (_FF, 17, "bold")
FONT_INPUT  = (_FF, 10)

THEMES = {
    "dark": dict(
        bg="#0f172a",    surface="#1e293b",  card="#1e293b",
        border="#334155", accent="#2563eb",  accent2="#10b981",
        text="#f1f5f9",  text2="#cbd5e1",   muted="#94a3b8",
        success="#10b981", error="#ef4444", warn="#f59e0b",
        btn_fg="#ffffff", warn_fg="#ffffff",
        scrollbar="#334155", card_ok_bg="#022c22",
        input_bg="#0f172a", log_bg="#020617", log_fg="#64748b",
    ),
    "light": dict(
        bg="#f1f5f9",    surface="#ffffff",  card="#ffffff",
        border="#e2e8f0", accent="#2563eb", accent2="#10b981",
        text="#0f172a",  text2="#334155",   muted="#64748b",
        success="#059669", error="#dc2626", warn="#d97706",
        btn_fg="#ffffff", warn_fg="#ffffff",
        scrollbar="#e2e8f0", card_ok_bg="#ecfdf5",
        input_bg="#f8fafc", log_bg="#0f172a", log_fg="#64748b",
    ),
}

LOG_ICONS = {"ok": "✓ ", "err": "✗ ", "warn": "⚠ ", "info": "ℹ ", "": "  "}

HELP_TEXT = """\
ESTRUCTURA ESPERADA DE ARCHIVOS

── Archivo People ──────────────────────────────
Columnas fijas (no son conceptos):
  CLAVE EMPLEADO · NOMBRE · AP. PATERNO · AP. MATERNO
  PUESTO · DEPARTAMENTO · SUCURSAL · TIPO DE PAGO
  FECHA INGRESO · NSS · RFC · CURP · EMISOR

Columnas de conceptos (dinámicas):
  Cualquier otra columna se trata como concepto.
  Ej: SUELDO, P2AH, BONO, etc.
  El importe es la SUMA de los conceptos seleccionados.

── Archivo Cuenta Vales ────────────────────────
Columnas requeridas:
  CLAVE EMPLEADO · CUENTA VALE

La unión se hace por CLAVE EMPLEADO (inner join).

── Formatos de dispersión TXT ──────────────────
Por Tarjeta  — Header 21 dígitos / Body 26 dígitos
  Header : cliente(7) + 0000(4) + total_centavos(10)
  Body   : cuenta_vale(16) + importe_centavos(10)

Por Empleado — Header 21 dígitos / Body 27 dígitos
  Header : empresa(7) + 0000(4) + total_centavos(10)
  Body   : departamento(7) + clave_empleado(10) + importe_centavos(10)
"""


# ─────────────────────────── Utilidades ────────────────────────────

def detect_system_theme() -> str:
    try:
        if platform.system() == "Darwin":
            r = subprocess.run(
                ["defaults", "read", "-g", "AppleInterfaceStyle"],
                capture_output=True, text=True, timeout=2,
            )
            return "dark" if "Dark" in r.stdout else "light"
        elif platform.system() == "Windows":
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Themes\Personalize",
            )
            val, _ = winreg.QueryValueEx(key, "AppsUseLightTheme")
            return "light" if val == 1 else "dark"
    except Exception:
        pass
    return "dark"


def validate_excel_file(path: str, header_row: int) -> dict:
    try:
        df = pd.read_excel(path, dtype=str, header=header_row - 1)
        orig_cols = [str(c) for c in df.columns]
        norm_cols = [c.strip().lower().replace(" ", "_") for c in orig_cols]
        has_key   = "clave_empleado" in norm_cols
        concept_cols = []
        for orig, norm in zip(orig_cols, norm_cols):
            if norm not in FIXED_PEOPLE_COLS:
                col_vals = pd.to_numeric(df[orig], errors="coerce")
                if (col_vals > 0).any():
                    concept_cols.append(orig)
        return {"ok": True, "rows": len(df), "has_key": has_key,
                "concept_cols": concept_cols}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def open_file(path: str):
    try:
        if platform.system() == "Darwin":
            subprocess.Popen(["open", path])
        elif platform.system() == "Windows":
            os.startfile(path)
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{e}")


# ─────────────────────────── Tooltip ────────────────────────────

class Tooltip:
    def __init__(self, widget, text: str):
        self._widget = widget
        self._text   = text
        self._win    = None
        widget.bind("<Enter>", self._show, add="+")
        widget.bind("<Leave>", self._hide, add="+")

    def _show(self, event=None):
        if self._win:
            return
        x = self._widget.winfo_rootx() + 10
        y = self._widget.winfo_rooty() + self._widget.winfo_height() + 6
        self._win = tw = tk.Toplevel(self._widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(tw, text=self._text, bg="#1e293b", fg="#f1f5f9",
                 font=FONT_SMALL, relief="flat", bd=1,
                 padx=8, pady=5, wraplength=340, justify="left").pack()

    def _hide(self, event=None):
        if self._win:
            self._win.destroy()
            self._win = None


# ─────────────────────────── TagInput ────────────────────────────

class TagInput(tk.Frame):
    """Input de conceptos con chips y dropdown filtrable."""

    def __init__(self, parent, T: dict, initial_tags: list | None = None, **kw):
        super().__init__(parent, bg=T["surface"], **kw)
        self._T                      = T
        self.tags: list[str]         = list(initial_tags or [])
        self._suggestions: list[str] = []
        self._dropdown: tk.Toplevel | None = None
        self._outside_click_id: str | None = None
        self._build()

    def _build(self):
        T = self._T

        # Chips area
        self._chips_frame = tk.Frame(self, bg=T["surface"])
        self._chips_frame.pack(fill="x", padx=0, pady=(0, 8))

        # Entry
        self._entry_wrap = tk.Frame(self, bg=T["input_bg"],
                                    highlightbackground=T["border"],
                                    highlightthickness=1)
        self._entry_wrap.pack(fill="x")

        self._entry = tk.Entry(
            self._entry_wrap, bg=T["input_bg"], fg=T["muted"],
            insertbackground=T["accent"],
            relief="flat", bd=0, font=FONT_INPUT, highlightthickness=0,
        )
        self._entry.pack(fill="x", padx=12, pady=9)
        self._ph = "Escribe el nombre del concepto y presiona Enter..."
        self._entry.insert(0, self._ph)

        self._entry.bind("<FocusIn>",    self._on_focus_in)
        self._entry.bind("<FocusOut>",   self._on_focus_out)
        self._entry.bind("<KeyRelease>", self._on_keyrelease)
        self._entry.bind("<Return>",     self._on_return)
        self._entry.bind("<Escape>",     lambda e: self._hide_dropdown())
        self._entry.bind("<BackSpace>",  self._on_backspace)

        self._rebuild_chips()

    # ── Focus / placeholder ───────────────────────────────────────

    def _on_focus_in(self, event=None):
        T = self._T
        if self._entry.get() == self._ph:
            self._entry.delete(0, "end")
            self._entry.config(fg=T["text"])
        self._entry_wrap.configure(highlightbackground=T["accent"])
        self._open_or_repopulate()

    def _on_focus_out(self, event=None):
        T = self._T
        self._entry_wrap.configure(highlightbackground=T["border"])
        if not self._entry.get().strip():
            self._entry.delete(0, "end")
            self._entry.insert(0, self._ph)
            self._entry.config(fg=T["muted"])
        self.after(150, self._maybe_hide)

    # ── Chips ─────────────────────────────────────────────────────

    def _rebuild_chips(self):
        T = self._T
        for w in self._chips_frame.winfo_children():
            w.destroy()
        for tag in self.tags:
            chip = tk.Frame(self._chips_frame, bg=T["accent"])
            chip.pack(side="left", padx=(0, 4), pady=2)
            tk.Label(chip, text=tag, bg=T["accent"], fg=T["btn_fg"],
                     font=(_FF, 9, "bold"), padx=8, pady=4).pack(side="left")
            tk.Button(chip, text="×",
                      command=lambda t=tag: self._remove_tag(t),
                      bg=T["accent"], fg=T["btn_fg"],
                      activebackground=T["accent"], activeforeground=T["btn_fg"],
                      font=(_FF, 11), relief="flat", bd=0,
                      highlightthickness=0, takefocus=0,
                      cursor="hand2", padx=4, pady=2).pack(side="left")

    def _remove_tag(self, tag):
        if tag in self.tags:
            self.tags.remove(tag)
        self._rebuild_chips()
        self._repopulate()

    def clear_all(self):
        self.tags.clear()
        self._rebuild_chips()
        self._repopulate()

    # ── Teclado ───────────────────────────────────────────────────

    def _on_return(self, event=None):
        text = self._entry.get().strip()
        if text and text != self._ph:
            tag = text.upper()
            if tag not in self.tags:
                self.tags.append(tag)
            self._entry.delete(0, "end")
            self._rebuild_chips()
            self._repopulate()

    def _on_backspace(self, event=None):
        val = self._entry.get()
        if (not val or val == self._ph) and self.tags:
            self.tags.pop()
            self._rebuild_chips()
            self._repopulate()

    def _on_keyrelease(self, event=None):
        if event and event.keysym in ("Return", "Escape", "Tab",
                                       "Up", "Down", "Left", "Right"):
            return
        self._open_or_repopulate()

    # ── Dropdown ──────────────────────────────────────────────────

    def _open_or_repopulate(self):
        if not self._suggestions:
            return
        if self._dropdown and self._dropdown.winfo_exists():
            self._repopulate()
        else:
            self._open_dropdown()

    def _on_outside_click(self, event):
        w = event.widget
        while w is not None:
            if w is self or (self._dropdown and
                             self._dropdown.winfo_exists() and
                             w is self._dropdown):
                return
            try:
                w = w.master
            except Exception:
                break
        self._hide_dropdown()

    def _open_dropdown(self):
        if not self._suggestions:
            return
        self.update_idletasks()
        T = self._T

        self._dropdown = dd = tk.Toplevel(self)
        dd.wm_overrideredirect(True)
        dd.configure(bg=T["border"])
        dd.bind("<FocusOut>", lambda e: self.after(150, self._maybe_hide))

        outer = tk.Frame(dd, bg=T["surface"])
        outer.pack(fill="both", expand=True, padx=1, pady=1)

        hdr = tk.Frame(outer, bg=T["surface"])
        hdr.pack(fill="x", padx=10, pady=(6, 4))
        self._hdr_lbl = tk.Label(hdr, text="Conceptos disponibles",
                                  bg=T["surface"], fg=T["muted"], font=FONT_SMALL)
        self._hdr_lbl.pack(side="left")
        tk.Frame(outer, bg=T["border"], height=1).pack(fill="x")

        lc = tk.Frame(outer, bg=T["surface"])
        lc.pack(fill="both", expand=True)
        self._canvas_dd = tk.Canvas(lc, bg=T["surface"], highlightthickness=0)
        vsb = tk.Scrollbar(lc, orient="vertical", command=self._canvas_dd.yview)
        self._canvas_dd.configure(yscrollcommand=vsb.set)
        self._list_frame = tk.Frame(self._canvas_dd, bg=T["surface"])
        self._list_win   = self._canvas_dd.create_window(
            (0, 0), window=self._list_frame, anchor="nw")
        self._list_frame.bind("<Configure>", lambda e: (
            self._canvas_dd.configure(scrollregion=self._canvas_dd.bbox("all")),
            self._canvas_dd.itemconfig(self._list_win,
                                        width=self._canvas_dd.winfo_width()),
        ))
        self._canvas_dd.bind("<Configure>", lambda e:
            self._canvas_dd.itemconfig(self._list_win, width=e.width))
        self._canvas_dd.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._repopulate()
        self._reposition()
        if not self._outside_click_id:
            self._outside_click_id = self.winfo_toplevel().bind(
                "<Button-1>", self._on_outside_click, add="+")

    def _repopulate(self):
        if not self._dropdown or not self._dropdown.winfo_exists():
            return
        T     = self._T
        query = self._entry.get().strip().upper()
        if query == self._ph.upper():
            query = ""
        tags_upper = {t.upper() for t in self.tags}
        visible = [s for s in self._suggestions if not query or query in s.upper()]

        self._hdr_lbl.config(
            text=f"Conceptos ({len(visible)})" if not query
            else f'"{self._entry.get().strip()}" — {len(visible)} resultado(s)')

        for w in self._list_frame.winfo_children():
            w.destroy()

        if not visible:
            tk.Label(self._list_frame, text="Sin resultados",
                     bg=T["surface"], fg=T["muted"], font=FONT_SMALL,
                     padx=12, pady=8).pack(anchor="w")
        else:
            for item in visible:
                self._make_row(item, item.upper() in tags_upper, T)
        self._reposition()

    def _make_row(self, item: str, selected: bool, T: dict):
        bg_n = T["surface"]
        bg_h = T["input_bg"]
        row  = tk.Frame(self._list_frame, bg=bg_n, cursor="hand2")
        row.pack(fill="x")
        bar  = tk.Frame(row, bg=T["accent"] if selected else bg_n, width=3)
        bar.pack(side="left", fill="y")
        lbl  = tk.Label(row, text=item, bg=bg_n,
                        fg=T["muted"] if selected else T["text"],
                        font=FONT_SMALL, anchor="w", padx=10, pady=6)
        lbl.pack(side="left", fill="x", expand=True)
        if not selected:
            def _e(e, r=row, l=lbl): r.configure(bg=bg_h); l.configure(bg=bg_h)
            def _l(e, r=row, l=lbl): r.configure(bg=bg_n); l.configure(bg=bg_n)
            row.bind("<Enter>", _e); row.bind("<Leave>", _l)
            lbl.bind("<Enter>", _e); lbl.bind("<Leave>", _l)
            row.bind("<Button-1>", lambda e, i=item: self._pick(i))
            lbl.bind("<Button-1>", lambda e, i=item: self._pick(i))

    def _pick(self, item: str):
        tag = item.strip().upper()
        if tag and tag not in self.tags:
            self.tags.append(tag)
        self._entry.delete(0, "end")
        self._rebuild_chips()
        self._hide_dropdown()
        self.winfo_toplevel().focus_set()

    def _reposition(self):
        if not self._dropdown or not self._dropdown.winfo_exists():
            return
        self.update_idletasks()
        ew = self._entry_wrap
        x  = ew.winfo_rootx()
        y  = ew.winfo_rooty() + ew.winfo_height()
        w  = ew.winfo_width()
        self._list_frame.update_idletasks()
        h  = min(250, self._list_frame.winfo_reqheight()) + 42
        self._canvas_dd.configure(height=max(h - 42, 40))
        self._dropdown.geometry(f"{w}x{h}+{x}+{y}")

    def _hide_dropdown(self):
        if self._dropdown and self._dropdown.winfo_exists():
            self._dropdown.destroy()
        self._dropdown = None
        if self._outside_click_id:
            try:
                self.winfo_toplevel().unbind("<Button-1>", self._outside_click_id)
            except Exception:
                pass
            self._outside_click_id = None

    def _maybe_hide(self):
        try:
            focused = str(self.focus_get())
            if focused == str(self._entry):
                return
            if self._dropdown and self._dropdown.winfo_exists():
                if focused.startswith(str(self._dropdown)):
                    return
        except Exception:
            pass
        self._hide_dropdown()

    def set_suggestions(self, concepts: list[str]):
        self._suggestions = list(concepts)
        self._repopulate()

    def get_tags(self) -> list[str]:
        return list(self.tags)


# ─────────────────────────── FilePickerCard ────────────────────────────

class FilePickerCard(tk.Frame):
    """Dropzone para archivo xlsx con borde discontinuo y fila de encabezado."""

    def __init__(self, parent, label: str, icon_idle: str, icon_ok: str,
                 T: dict, path_var: tk.StringVar,
                 header_row_var: tk.StringVar | None = None,
                 on_concepts_found=None, **kw):
        super().__init__(parent, bg=T["surface"], **kw)
        self._T                 = T
        self._label             = label
        self._icon_idle         = icon_idle
        self._icon_ok           = icon_ok
        self.path_var           = path_var
        self._header_row_var    = header_row_var
        self._on_concepts_found = on_concepts_found
        self._loaded            = False
        self._build()

    def _build(self):
        T  = self._T
        bg = T["surface"]

        self._cv = tk.Canvas(self, bg=bg, highlightthickness=0, cursor="hand2")
        self._cv.pack(fill="both", expand=True)
        self._cv.bind("<Configure>", self._redraw_border)
        self._cv.bind("<Button-1>",  lambda e: self._pick())

        self._inner = tk.Frame(self._cv, bg=bg, cursor="hand2")
        self._win   = self._cv.create_window((0, 0), window=self._inner, anchor="nw")

        # Status dot
        self._dot = tk.Label(self._inner, text="●", bg=bg, fg=T["muted"],
                             font=(_FF, 9))
        self._dot.place(relx=1.0, rely=0.0, x=-10, y=8, anchor="ne")

        # Center content
        center = tk.Frame(self._inner, bg=bg)
        center.place(relx=0.5, rely=0.40, anchor="center")

        self._icon_lbl = tk.Label(center, text=self._icon_idle, bg=bg,
                                   fg=T["muted"], font=(_FF, 26))
        self._icon_lbl.pack(pady=(0, 6))
        self._title_lbl = tk.Label(center, text=self._label, bg=bg,
                                    fg=T["text"], font=FONT_BADGE)
        self._title_lbl.pack()
        self._name_lbl = tk.Label(center, text="Arrastra archivo o examina...",
                                   bg=bg, fg=T["muted"], font=FONT_SMALL,
                                   wraplength=160)
        self._name_lbl.pack(pady=(4, 0))

        # Bottom: header row spinbox
        bottom = tk.Frame(self._inner, bg=bg)
        bottom.place(relx=0.5, rely=0.88, anchor="center")
        if self._header_row_var is not None:
            tk.Label(bottom, text="FILA ENCABEZADO:",
                     bg=bg, fg=T["muted"], font=(_FF, 8, "bold")).pack(side="left")
            spb = tk.Spinbox(
                bottom, textvariable=self._header_row_var,
                from_=1, to=100, width=3,
                bg=bg, fg=T["text"], insertbackground=T["accent"],
                relief="flat", bd=0, font=FONT_INPUT,
                buttonbackground=bg, highlightthickness=0,
            )
            spb.pack(side="left", padx=(6, 0))
            bottom.bind("<Button-1>", lambda e: "break")

        for w in (self._inner, center, self._icon_lbl,
                  self._title_lbl, self._name_lbl):
            w.bind("<Button-1>", lambda e: self._pick())

        if HAS_DND:
            self._cv.drop_target_register(DND_FILES)
            self._cv.dnd_bind("<<Drop>>",     self._on_drop)
            self._cv.dnd_bind("<<DragEnter>>", self._on_drag_enter)
            self._cv.dnd_bind("<<DragLeave>>", self._on_drag_leave)

    # ── Border drawing ────────────────────────────────────────────

    def _redraw_border(self, event=None, color=None, width=1):
        T = self._T
        w = self._cv.winfo_width()
        h = self._cv.winfo_height()
        self._cv.itemconfig(self._win, width=w, height=h)
        self._cv.delete("db")
        c = color or T["border"]
        self._cv.create_rectangle(3, 3, w-3, h-3,
                                   outline=c, dash=(6, 3), width=width, tags="db")

    def _on_drag_enter(self, event=None):
        self._redraw_border(color=self._T["accent"], width=2)

    def _on_drag_leave(self, event=None):
        c = self._T["accent2"] if self._loaded else None
        self._redraw_border(color=c, width=2 if self._loaded else 1)

    def _on_drop(self, event):
        self._on_drag_leave()
        path = event.data.strip().strip("{}")
        if os.path.isfile(path) and path.lower().endswith((".xlsx", ".xls")):
            self._set_path(path)

    # ── File selection ────────────────────────────────────────────

    def _pick(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel Files", "*.xlsx *.xls"), ("All", "*.*")])
        if path:
            self._set_path(path)

    def _set_path(self, path):
        self.path_var.set(path)
        self._name_lbl.config(text="Validando…", fg=self._T["muted"])
        try:
            header_row = int(self._header_row_var.get()) if self._header_row_var else 1
        except (ValueError, AttributeError):
            header_row = 1

        def worker():
            result = validate_excel_file(path, header_row)
            self.after(0, lambda: self._show_validation(result, path))
        threading.Thread(target=worker, daemon=True).start()

    def _show_validation(self, result: dict, path: str):
        T     = self._T
        fname = os.path.basename(path)
        display = (fname[:20] + "…") if len(fname) > 20 else fname

        if not result["ok"]:
            self._name_lbl.config(text=result["error"][:32], fg=T["error"])
            self._dot.config(fg=T["error"])
            return

        self._loaded = True
        ok_bg = T["card_ok_bg"]
        for w in self._all_widgets(self._inner):
            try:
                w.configure(bg=ok_bg)
            except tk.TclError:
                pass
        self._cv.configure(bg=ok_bg)

        self._redraw_border(color=T["accent2"], width=2)
        self._dot.configure(fg=T["success"])
        self._icon_lbl.configure(text=self._icon_ok, fg=T["accent2"])
        self._name_lbl.configure(text=display, fg=T["text2"])

        if self._on_concepts_found and result.get("concept_cols"):
            self._on_concepts_found(result["concept_cols"])

    def _all_widgets(self, w):
        yield w
        for child in w.winfo_children():
            yield from self._all_widgets(child)

    def get_path(self) -> str:
        return self.path_var.get()


# ─────────────────────────── BitacoraPanel ────────────────────────────

class BitacoraPanel(tk.Frame):
    """Consola de terminal para logs de proceso."""

    def __init__(self, parent, T: dict, **kw):
        super().__init__(parent, bg=T["surface"],
                         highlightbackground=T["border"],
                         highlightthickness=1, **kw)
        self._T = T
        self._build()

    def _build(self):
        T = self._T
        hdr = tk.Frame(self, bg=T["surface"])
        hdr.pack(fill="x", padx=14, pady=(10, 0))

        tk.Label(hdr, text="▶_  BITÁCORA DE EVENTOS",
                 bg=T["surface"], fg=T["muted"],
                 font=(_FF, 9, "bold")).pack(side="left")

        tk.Button(hdr, text="⌫", command=self.clear,
                  bg=T["surface"], fg=T["muted"],
                  activebackground=T["surface"], activeforeground=T["error"],
                  font=(_FF, 13), relief="flat", bd=0,
                  highlightthickness=0, takefocus=0,
                  cursor="hand2").pack(side="right")

        self.text = tk.Text(
            self, bg=T["log_bg"], fg=T["log_fg"],
            font=FONT_MONO, relief="flat", bd=0,
            state="disabled", wrap="word",
            selectbackground=T["accent"],
            padx=12, pady=8,
        )
        self.text.pack(fill="both", expand=True, padx=6, pady=(4, 6))

        self.text.tag_config("ok",   foreground="#10b981")
        self.text.tag_config("err",  foreground="#ef4444")
        self.text.tag_config("warn", foreground="#f59e0b")
        self.text.tag_config("info", foreground="#60a5fa")
        self.text.tag_config("ts",   foreground="#475569")

    def log(self, msg: str, kind: str = ""):
        self.text.config(state="normal")
        ts   = datetime.datetime.now().strftime("%H:%M:%S")
        icon = LOG_ICONS.get(kind, "  ")
        self.text.insert("end", f"[{ts}] ", "ts")
        self.text.insert("end", icon + msg + "\n", kind)
        self.text.see("end")
        self.text.config(state="disabled")

    def clear(self):
        self.text.config(state="normal")
        self.text.delete("1.0", "end")
        self.text.config(state="disabled")

    def copy(self):
        content = self.text.get("1.0", "end-1c")
        self.clipboard_clear()
        self.clipboard_append(content)


# ─────────────────────────── Generadores Omonel ────────────────────────────

def generate_omonel_layout(result_df: pd.DataFrame, output_path: str) -> str:
    """Genera el layout de dispersión bancaria Omonel en formato xlsx."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dispersión Omonel"

    thin       = Side(style="thin", color="D0D7DE")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill   = PatternFill("solid", fgColor="2D3748")
    hdr_font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="left", vertical="center")
    num_align  = Alignment(horizontal="right", vertical="center")
    total_fill = PatternFill("solid", fgColor="D5EDE9")
    total_font = Font(name="Arial", bold=True, size=10)

    ws.merge_cells("A1:H1")
    ws["A1"] = "OMONEL  |  LAYOUT DE DISPERSIÓN BANCARIA"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="2D3748")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ts = datetime.datetime.now().strftime("%d/%m/%Y  %H:%M:%S")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Generado: {ts}"
    ws["A2"].font      = Font(name="Arial", italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[3].height = 6

    HEADERS = ["NO.", "CLAVE EMPLEADO", "RFC", "NOMBRE COMPLETO",
               "CLABE INTERBANCARIA", "BANCO", "CUENTA VALE", "IMPORTE ($)"]
    for col, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[4].height = 28

    alt_fill = PatternFill("solid", fgColor="EBF5F3")
    row_num  = 5
    for i, rec in enumerate(result_df.itertuples(index=False), 1):
        fill = alt_fill if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        values = [i,
                  getattr(rec, "clave_empleado", ""),
                  getattr(rec, "rfc", ""),
                  getattr(rec, "nombre", ""),
                  getattr(rec, "clabe", ""),
                  getattr(rec, "banco", ""),
                  getattr(rec, "cuenta_vale", ""),
                  getattr(rec, "importe", 0)]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = data_font; cell.fill = fill; cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 8:
                cell.alignment = num_align
                cell.number_format = '"$"#,##0.00'
            else:
                cell.alignment = data_align
        ws.row_dimensions[row_num].height = 18
        row_num += 1

    ws.cell(row=row_num, column=7, value="TOTAL DISPERSIÓN:").font  = total_font
    ws.cell(row=row_num, column=7).fill      = total_fill
    ws.cell(row=row_num, column=7).alignment = Alignment(horizontal="right")
    ws.cell(row=row_num, column=7).border    = border
    total_cell = ws.cell(row=row_num, column=8, value=f"=SUM(H5:H{row_num-1})")
    total_cell.font = total_font; total_cell.fill = total_fill
    total_cell.alignment = num_align
    total_cell.number_format = '"$"#,##0.00'; total_cell.border = border
    ws.row_dimensions[row_num].height = 22

    for i, w in enumerate([6, 18, 16, 32, 22, 18, 18, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Resumen")
    ws2["A1"] = "RESUMEN DE DISPERSIÓN"
    ws2["A1"].font = Font(name="Arial", bold=True, size=13, color="2D3748")
    ws2.append([])
    ws2.append(["Total de registros", len(result_df)])
    total_imp = result_df["importe"].sum() if "importe" in result_df.columns else 0
    ws2.append(["Importe total a dispersar", f"${total_imp:,.2f}"])
    ws2.append(["Fecha de generación", ts])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 22

    wb.save(output_path)
    return output_path


def generate_omonel_txt(result_df: pd.DataFrame, output_path: str,
                        numero_cliente: str) -> str:
    """Header 21 / Body 26  — dispersión por tarjeta."""
    num   = numero_cliente.strip().zfill(7)[:7]
    total = int(round(result_df["importe"].sum() * 100))
    lines = [f"{num}0000{total:010d}"]
    for rec in result_df.itertuples(index=False):
        cuenta = str(getattr(rec, "cuenta_vale", "")).strip().zfill(16)[:16]
        imp    = int(round(getattr(rec, "importe", 0) * 100))
        lines.append(f"{cuenta}{imp:010d}")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return output_path


def generate_omonel_txt_empleado(result_df: pd.DataFrame, output_path: str,
                                  numero_empresa: str) -> str:
    """Header 21 / Body 27  — dispersión por número de empleado."""
    num   = numero_empresa.strip().zfill(7)[:7]
    total = int(round(result_df["importe"].sum() * 100))
    lines = [f"{num}0000{total:010d}"]
    for rec in result_df.itertuples(index=False):
        depto_raw    = str(getattr(rec, "departamento", "")).strip()
        depto_digits = "".join(filter(str.isdigit, depto_raw))
        depto        = depto_digits.zfill(7)[:7] if depto_digits else "0000000"
        empleado     = str(getattr(rec, "clave_empleado", "")).strip().zfill(10)[:10]
        imp          = int(round(getattr(rec, "importe", 0) * 100))
        lines.append(f"{depto}{empleado}{imp:010d}")
    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return output_path


# ──────────────────────────── Lógica de negocio ───────────────────────────

FIXED_PEOPLE_COLS = {
    "clave_empleado", "nombre", "ap._paterno", "ap._materno",
    "puesto", "departamento", "sucursal", "tipo_de_pago",
    "fecha_ingreso", "nss", "rfc", "curp", "emisor",
}


def process_files(
    path_people:       str,
    path_vales:        str,
    conceptos:         list[str],
    log_fn,
    progress_fn,
    header_row_people: int = 1,
    header_row_vales:  int = 1,
) -> dict | None:
    """
    Lee ambos archivos, cruza por clave_empleado y calcula importes.
    Retorna {"df": DataFrame, "unmatched": int} o None si hay error.
    """
    try:
        progress_fn(0)
        log_fn(f"Leyendo People… (fila de encabezados: {header_row_people})", "info")
        df_people = pd.read_excel(path_people, dtype=str, header=header_row_people - 1)
        df_people.columns = [c.strip().lower().replace(" ", "_") for c in df_people.columns]
        log_fn(f"  ↳ {len(df_people):,} filas | columnas: {list(df_people.columns)}", "")

        log_fn(f"Leyendo Cuenta Vales… (fila de encabezados: {header_row_vales})", "info")
        df_vales = pd.read_excel(path_vales, dtype=str, header=header_row_vales - 1)
        df_vales.columns = [c.strip().lower().replace(" ", "_") for c in df_vales.columns]
        log_fn(f"  ↳ {len(df_vales):,} filas | columnas: {list(df_vales.columns)}", "")

        for label, df in [("People", df_people), ("Cuenta Vales", df_vales)]:
            if "clave_empleado" not in df.columns:
                log_fn(f"ERROR: '{label}' no tiene columna 'clave_empleado'", "err")
                return None

        df_people["clave_empleado"] = df_people["clave_empleado"].str.strip().str.upper()
        df_vales["clave_empleado"]  = df_vales["clave_empleado"].str.strip().str.upper()

        concept_cols = [c for c in df_people.columns if c not in FIXED_PEOPLE_COLS]
        log_fn(f"Conceptos detectados: {concept_cols}", "info")

        if conceptos:
            conceptos_norm = [c.strip().lower().replace(" ", "_") for c in conceptos]
            selected_cols  = [c for c in concept_cols if c in conceptos_norm]
            missing        = [c for c in conceptos_norm if c not in selected_cols]
            if missing:
                log_fn(f"Sin columna para: {missing}", "warn")
            if not selected_cols:
                log_fn("ERROR: Ningún concepto encontrado como columna.", "err")
                return None
            log_fn(f"Conceptos seleccionados: {selected_cols}", "info")
        else:
            selected_cols = concept_cols
            log_fn("Sin filtro: se suman todas las columnas de concepto.", "warn")

        df_people["importe"] = (
            df_people[selected_cols]
            .apply(pd.to_numeric, errors="coerce")
            .fillna(0)
            .sum(axis=1)
        )

        progress_fn(1)
        total_people = len(df_people)
        col_cv = "cuenta_vale" if "cuenta_vale" in df_vales.columns else df_vales.columns[-1]
        df_merge = df_people.merge(
            df_vales[["clave_empleado", col_cv]],
            on="clave_empleado", how="inner",
        )
        unmatched = total_people - len(df_merge)
        log_fn(f"Coincidencias: {len(df_merge):,}  |  Sin coincidencia: {unmatched:,}", "ok")

        if len(df_merge) == 0:
            log_fn("No se encontraron coincidencias.", "warn")
            return None

        def _col(df, *names, default=""):
            for n in names:
                if n in df.columns:
                    return df[n].fillna("").str.strip()
            return pd.Series([default] * len(df), index=df.index)

        nombre_completo = (
            _col(df_merge, "nombre") + " " +
            _col(df_merge, "ap._paterno") + " " +
            _col(df_merge, "ap._materno")
        ).str.strip()

        out = pd.DataFrame()
        out["clave_empleado"] = df_merge["clave_empleado"]
        out["rfc"]            = _col(df_merge, "rfc")
        out["nombre"]         = nombre_completo
        out["clabe"]          = _col(df_merge, "clabe", "clabe_interbancaria")
        out["banco"]          = _col(df_merge, "banco")
        out["departamento"]   = _col(df_merge, "departamento")
        out["cuenta_vale"]    = df_merge[col_cv]
        out["importe"]        = df_merge["importe"]

        return {"df": out, "unmatched": unmatched}

    except Exception as e:
        log_fn(f"Error procesando archivos: {e}", "err")
        return None


# ─────────────────────────── Ventana principal ────────────────────────────

if HAS_DND:
    _BaseApp = TkinterDnD.Tk
else:
    _BaseApp = tk.Tk


class OmonelApp(_BaseApp):
    def __init__(self):
        super().__init__()
        self.title("Dispersión Omonel  |  Generador de Layout")
        self.geometry("1040x760")
        self.minsize(880, 640)

        self._theme = detect_system_theme()
        self._T     = THEMES[self._theme]
        self.configure(bg=self._T["bg"])

        # Suppress macOS native button rendering
        self.option_add("*Button.relief",             "flat")
        self.option_add("*Button.borderWidth",        0)
        self.option_add("*Button.highlightThickness", 0)
        self.option_add("*Button.takeFocus",          0)

        self._people_path      = tk.StringVar()
        self._vales_path       = tk.StringVar()
        self._header_row       = tk.StringVar(value="7")
        self._header_row_vales = tk.StringVar(value="1")
        self._tipo             = tk.StringVar(value="empleado")
        self._cliente          = tk.StringVar()
        self._out_path         = tk.StringVar(value=os.path.join(
            os.path.expanduser("~"), "layout_omonel.txt"))
        self._tags: list[str]  = []

        self._queue:   queue.Queue = queue.Queue()
        self._working: bool        = False

        self._build_ui()

    # ──────────────────────────────────────────────────────────────────
    # HELPERS
    # ──────────────────────────────────────────────────────────────────

    def _card(self, parent, **kw) -> tk.Frame:
        T = self._T
        return tk.Frame(parent, bg=T["surface"],
                        highlightbackground=T["border"],
                        highlightthickness=1, **kw)

    def _sec_hdr(self, parent, num: str, title: str):
        T   = self._T
        hdr = tk.Frame(parent, bg=T["surface"])
        hdr.pack(fill="x", padx=16, pady=(14, 12))
        tk.Label(hdr, text=num, bg=T["accent"], fg=T["btn_fg"],
                 font=FONT_BADGE, padx=8, pady=3).pack(side="left")
        tk.Label(hdr, text=title, bg=T["surface"], fg=T["text"],
                 font=FONT_LABEL).pack(side="left", padx=10)
        return hdr

    def _input_field(self, parent, textvariable, icon="⚙",
                     placeholder="", editable=True) -> tk.Frame:
        T   = self._T
        box = tk.Frame(parent, bg=T["input_bg"],
                       highlightbackground=T["border"], highlightthickness=1)
        if icon:
            tk.Label(box, text=icon, bg=T["input_bg"], fg=T["muted"],
                     font=(_FF, 11), padx=8).pack(side="left")
            tk.Frame(box, bg=T["border"], width=1).pack(side="left", fill="y", pady=6)
        state = "normal" if editable else "readonly"
        e = tk.Entry(box, textvariable=textvariable, width=20,
                     bg=T["input_bg"], fg=T["text"],
                     readonlybackground=T["input_bg"],
                     insertbackground=T["accent"],
                     relief="flat", bd=0, font=FONT_INPUT,
                     highlightthickness=0, state=state)
        e.pack(side="left", fill="x", expand=True, padx=8, ipady=7)
        e.bind("<FocusIn>",  lambda ev: box.configure(highlightbackground=T["accent"]))
        e.bind("<FocusOut>", lambda ev: box.configure(highlightbackground=T["border"]))
        return box

    def _btn(self, parent, text, command, bg_key="accent",
             fg_key="btn_fg", **kw) -> tk.Button:
        T = self._T
        bg = T[bg_key]; fg = T[fg_key]
        b  = tk.Button(parent, text=text, command=command,
                       bg=bg, fg=fg,
                       activebackground=bg, activeforeground=fg,
                       font=FONT_BTN, relief="flat", bd=0,
                       highlightthickness=0, takefocus=0,
                       cursor="hand2", **kw)
        return b

    def _divider(self, parent):
        T = self._T
        tk.Frame(parent, bg=T["border"], height=1).pack(fill="x")

    # ──────────────────────────────────────────────────────────────────
    # UI BUILD
    # ──────────────────────────────────────────────────────────────────

    def _build_ui(self):
        T = self._T

        # ── HEADER ───────────────────────────────────────────────────
        header = tk.Frame(self, bg=T["surface"],
                          highlightbackground=T["border"], highlightthickness=1)
        header.pack(fill="x", side="top")

        inner_h = tk.Frame(header, bg=T["surface"])
        inner_h.pack(fill="x", padx=20, pady=10)

        # Logo
        logo = tk.Frame(inner_h, bg=T["accent"], width=40, height=40)
        logo.pack(side="left"); logo.pack_propagate(False)
        tk.Label(logo, text="⊟", bg=T["accent"], fg=T["btn_fg"],
                 font=(_FF, 18, "bold")).place(relx=0.5, rely=0.5, anchor="center")

        title_block = tk.Frame(inner_h, bg=T["surface"])
        title_block.pack(side="left", padx=12)
        tk.Label(title_block, text="DISPERSIÓN OMONEL",
                 bg=T["surface"], fg=T["text"],
                 font=FONT_TITLE).pack(anchor="w")
        tk.Label(title_block,
                 text=f"Generador de Layout Bancario  //  v{APP_VERSION}",
                 bg=T["surface"], fg=T["muted"], font=FONT_SMALL).pack(anchor="w")

        # Right controls
        help_btn = tk.Button(
            inner_h, text="ⓘ  Manual de Ayuda",
            command=self._show_help,
            bg=T["accent"], fg=T["btn_fg"],
            activebackground=T["accent"], activeforeground=T["btn_fg"],
            font=(_FF, 9, "bold"), relief="flat", bd=0,
            highlightthickness=0, takefocus=0,
            cursor="hand2", padx=14, pady=7,
        )
        help_btn.pack(side="right")

        theme_icon = "☾" if self._theme == "light" else "☀"
        theme_btn = tk.Button(
            inner_h, text=theme_icon,
            command=self._toggle_theme,
            bg=T["surface"], fg=T["muted"],
            activebackground=T["surface"], activeforeground=T["text"],
            font=(_FF, 15), relief="flat", bd=0,
            highlightthickness=1, highlightbackground=T["border"],
            takefocus=0, cursor="hand2",
            padx=8, pady=5,
        )
        theme_btn.pack(side="right", padx=(0, 10))

        # ── TWO-COLUMN BODY ───────────────────────────────────────────
        body = tk.Frame(self, bg=T["bg"])
        body.pack(fill="both", expand=True, padx=16, pady=14)
        body.columnconfigure(0, weight=55, minsize=440)
        body.columnconfigure(1, weight=45, minsize=300)
        body.rowconfigure(0, weight=1)

        # ── LEFT COLUMN ───────────────────────────────────────────────
        left = tk.Frame(body, bg=T["bg"])
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left.rowconfigure(2, weight=1)
        left.columnconfigure(0, weight=1)

        # Section 01 — Archivos de entrada
        sec01 = self._card(left)
        sec01.grid(row=0, column=0, sticky="ew", pady=(0, 8))

        hdr01 = self._sec_hdr(sec01, "01", "ARCHIVOS DE ENTRADA")

        cards_row = tk.Frame(sec01, bg=T["surface"])
        cards_row.pack(fill="x", padx=14, pady=(0, 14))
        cards_row.columnconfigure(0, weight=1)
        cards_row.columnconfigure(1, weight=1)

        self.picker_people = FilePickerCard(
            cards_row, label="ARCHIVO PEOPLE",
            icon_idle="⬆", icon_ok="✓", T=T,
            path_var=self._people_path,
            header_row_var=self._header_row,
            on_concepts_found=lambda cols: self.tag_input.set_suggestions(cols),
        )
        self.picker_people.grid(row=0, column=0, sticky="nsew",
                                padx=(0, 6), ipady=30)

        self.picker_vales = FilePickerCard(
            cards_row, label="CUENTA VALES",
            icon_idle="⬆", icon_ok="✓", T=T,
            path_var=self._vales_path,
            header_row_var=self._header_row_vales,
        )
        self.picker_vales.grid(row=0, column=1, sticky="nsew",
                               padx=(6, 0), ipady=30)

        # Section 02 — Conceptos
        sec02 = self._card(left)
        sec02.grid(row=1, column=0, sticky="ew", pady=(0, 8))

        hdr02 = self._sec_hdr(sec02, "02", "CONCEPTOS")
        # "ELIMINAR TODO" en el mismo header
        elim_btn = tk.Button(
            hdr02, text="🗑  ELIMINAR TODO",
            command=lambda: self.tag_input.clear_all(),
            bg=T["surface"], fg=T["error"],
            activebackground=T["surface"], activeforeground=T["error"],
            font=(_FF, 8, "bold"), relief="flat", bd=0,
            highlightthickness=0, takefocus=0, cursor="hand2",
        )
        elim_btn.pack(side="right")

        self.tag_input = TagInput(sec02, T=T, initial_tags=self._tags)
        self.tag_input.pack(fill="x", padx=14, pady=(0, 14))

        # Section 04 — Bitácora (expands)
        self.log_panel = BitacoraPanel(left, T=T)
        self.log_panel.grid(row=2, column=0, sticky="nsew")
        self.log_panel.log("Sistema listo. Carga los archivos y presiona Generar.", "info")

        # ── RIGHT COLUMN — Section 03 ─────────────────────────────────
        right = tk.Frame(body, bg=T["bg"])
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        sec03 = self._card(right)
        sec03.grid(row=0, column=0, sticky="nsew")
        sec03.rowconfigure(0, weight=1)
        sec03.columnconfigure(0, weight=1)

        # Scrollable inner for sec03
        s3_canvas = tk.Canvas(sec03, bg=T["surface"], highlightthickness=0)
        s3_vsb    = tk.Scrollbar(sec03, orient="vertical", command=s3_canvas.yview)
        s3_canvas.configure(yscrollcommand=s3_vsb.set)
        s3_inner  = tk.Frame(s3_canvas, bg=T["surface"])
        s3_win    = s3_canvas.create_window((0, 0), window=s3_inner, anchor="nw")

        def _s3conf(e):
            s3_canvas.configure(scrollregion=s3_canvas.bbox("all"))
            s3_canvas.itemconfig(s3_win, width=s3_canvas.winfo_width())
        s3_inner.bind("<Configure>", _s3conf)
        s3_canvas.bind("<Configure>", lambda e:
            s3_canvas.itemconfig(s3_win, width=e.width))

        s3_canvas.grid(row=0, column=0, sticky="nsew")
        s3_vsb.grid(row=0, column=1, sticky="ns")

        s3 = tk.Frame(s3_inner, bg=T["surface"])
        s3.pack(fill="both", expand=True, padx=18, pady=16)

        # Header sec03
        hdr03 = tk.Frame(s3, bg=T["surface"])
        hdr03.pack(fill="x", pady=(0, 16))
        tk.Label(hdr03, text="03", bg=T["accent"], fg=T["btn_fg"],
                 font=FONT_BADGE, padx=8, pady=3).pack(side="left")
        tk.Label(hdr03, text="ARCHIVO DE SALIDA", bg=T["surface"], fg=T["text"],
                 font=FONT_LABEL).pack(side="left", padx=10)

        # ── Tipo de Dispersión ────────────────────────────────────────
        tk.Label(s3, text="TIPO DE DISPERSIÓN", bg=T["surface"], fg=T["muted"],
                 font=(_FF, 8, "bold")).pack(anchor="w", pady=(0, 8))

        self._tipo_cards: dict = {}
        for val, icon_char, lbl_text in [
            ("empleado", "◉", "Por No. Empleado"),
            ("tarjeta",  "▤", "Por Tarjeta"),
        ]:
            c_outer = tk.Frame(s3, bg=T["border"])
            c_outer.pack(fill="x", pady=(0, 4))
            c_card  = tk.Frame(c_outer, bg=T["card"], cursor="hand2")
            c_card.pack(fill="both", expand=True, padx=1, pady=1)
            c_inner = tk.Frame(c_card, bg=T["card"])
            c_inner.pack(fill="x", padx=12, pady=10)

            ic_lbl = tk.Label(c_inner, text=icon_char, bg=T["card"],
                              fg=T["muted"], font=(_FF, 14))
            ic_lbl.pack(side="left", padx=(0, 10))

            txt_box = tk.Frame(c_inner, bg=T["card"])
            txt_box.pack(side="left", fill="x", expand=True)
            main_lbl = tk.Label(txt_box, text=lbl_text, bg=T["card"],
                                fg=T["text"], font=(_FF, 10, "bold"))
            main_lbl.pack(anchor="w")
            sub_lbl = tk.Label(txt_box, text="Selección actual",
                               bg=T["card"], fg=T["muted"], font=FONT_SMALL)
            sub_lbl.pack(anchor="w")

            chk_lbl = tk.Label(c_inner, text="", bg=T["card"],
                               fg=T["btn_fg"], font=(_FF, 14))
            chk_lbl.pack(side="right")

            all_w = [c_outer, c_card, c_inner, ic_lbl,
                     txt_box, main_lbl, sub_lbl, chk_lbl]
            self._tipo_cards[val] = {
                "outer": c_outer, "card": c_card, "inner": c_inner,
                "icon": ic_lbl, "txt_box": txt_box,
                "main": main_lbl, "sub": sub_lbl, "check": chk_lbl,
                "all": all_w,
            }
            for w in [c_outer, c_card, c_inner, ic_lbl,
                      main_lbl, sub_lbl, txt_box]:
                w.bind("<Button-1>", lambda e, v=val: self._select_tipo(v))

        # ── Número empresa ────────────────────────────────────────────
        self.cli_label = tk.Label(s3, text="NÚMERO DE EMPRESA",
                                   bg=T["surface"], fg=T["muted"],
                                   font=(_FF, 8, "bold"))
        self.cli_label.pack(anchor="w", pady=(14, 6))

        num_field = self._input_field(s3, self._cliente, icon="⚙")
        num_field.pack(fill="x", pady=(0, 14))

        # ── GENERAR button ────────────────────────────────────────────
        self.btn_run = tk.Button(
            s3, text="GENERAR LAYOUT OMONEL  →",
            command=self._run,
            bg=T["accent"], fg=T["btn_fg"],
            activebackground=T["accent"], activeforeground=T["btn_fg"],
            font=(_FF, 10, "bold"), relief="flat", bd=0,
            highlightthickness=0, takefocus=0, cursor="hand2",
        )
        self.btn_run.pack(fill="x", ipady=13, pady=(0, 16))

        # ── Summary metrics (always visible, updated post-gen) ────────
        self._divider(s3)
        metrics_row = tk.Frame(s3, bg=T["surface"])
        metrics_row.pack(fill="x", pady=14)
        metrics_row.columnconfigure(0, weight=1)
        metrics_row.columnconfigure(1, weight=1)
        metrics_row.columnconfigure(2, weight=1)

        self._metric_vars = {}
        for col, (key, lbl_text) in enumerate([
            ("registros", "REGISTROS"),
            ("importe",   "IMPORTE"),
            ("sin_match", "SIN MATCH"),
        ]):
            cell = tk.Frame(metrics_row, bg=T["surface"])
            cell.grid(row=0, column=col, sticky="ew",
                      padx=(0, 8) if col < 2 else 0)
            tk.Label(cell, text=lbl_text, bg=T["surface"], fg=T["muted"],
                     font=(_FF, 8, "bold")).pack(anchor="w")
            var = tk.StringVar(value="—")
            lbl = tk.Label(cell, textvariable=var, bg=T["surface"],
                           fg=T["text"], font=FONT_NUM)
            lbl.pack(anchor="w")
            self._metric_vars[key] = var

        # ── File action buttons ───────────────────────────────────────
        self._divider(s3)
        self._file_btns_frame = tk.Frame(s3, bg=T["surface"])
        self._file_btns_frame.pack(fill="x", pady=(12, 0))

        tk.Label(self._file_btns_frame, text="ABRIR LAYOUT GENERADO EN:",
                 bg=T["surface"], fg=T["accent"],
                 font=(_FF, 8, "bold")).pack(anchor="w", pady=(0, 8))

        file_row = tk.Frame(self._file_btns_frame, bg=T["surface"])
        file_row.pack(fill="x")
        file_row.columnconfigure(0, weight=1)
        file_row.columnconfigure(1, weight=1)

        self._btn_txt_card  = self._file_btn_card(file_row, "📄", "FORMATO .TXT",  None)
        self._btn_txt_card.grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self._btn_xlsx_card = self._file_btn_card(file_row, "⊞", "FORMATO .XLSX", None)
        self._btn_xlsx_card.grid(row=0, column=1, sticky="ew")

        # ── Ruta de guardado ──────────────────────────────────────────
        self._divider(s3)
        path_row = tk.Frame(s3, bg=T["surface"])
        path_row.pack(fill="x", pady=(10, 4))

        tk.Label(path_row, text="RUTA DE GUARDADO ACTUAL:",
                 bg=T["surface"], fg=T["muted"],
                 font=(_FF, 8, "bold")).pack(side="left")
        tk.Button(path_row, text="CAMBIAR",
                  command=self._pick_output,
                  bg=T["surface"], fg=T["accent"],
                  activebackground=T["surface"], activeforeground=T["accent"],
                  font=(_FF, 8, "bold"), relief="flat", bd=0,
                  highlightthickness=0, takefocus=0,
                  cursor="hand2").pack(side="right")

        self._path_lbl = tk.Label(s3, textvariable=self._out_path,
                                   bg=T["surface"], fg=T["muted"],
                                   font=(_FF, 8), wraplength=260,
                                   justify="left", anchor="w")
        self._path_lbl.pack(fill="x", pady=(0, 14))

        # Init tipo
        self._select_tipo(self._tipo.get())

    def _file_btn_card(self, parent, icon: str, label: str,
                       cmd) -> tk.Frame:
        T = self._T
        outer = tk.Frame(parent, bg=T["border"])
        card  = tk.Frame(outer, bg=T["surface"], cursor="hand2")
        card.pack(fill="both", expand=True, padx=1, pady=1)
        inner = tk.Frame(card, bg=T["surface"])
        inner.pack(expand=True, pady=12)
        tk.Label(inner, text=icon, bg=T["surface"], fg=T["muted"],
                 font=(_FF, 20)).pack()
        tk.Label(inner, text=label, bg=T["surface"], fg=T["text"],
                 font=(_FF, 8, "bold")).pack(pady=(4, 0))
        if cmd:
            for w in [outer, card, inner]:
                w.bind("<Button-1>", lambda e: cmd())
        outer._card  = card
        outer._inner = inner
        outer._cmd   = cmd
        return outer

    def _update_file_btn_card(self, card_outer: tk.Frame, cmd):
        T = self._T
        card_outer._cmd = cmd
        for w in [card_outer, card_outer._card, card_outer._inner]:
            w.bind("<Button-1>", lambda e, c=cmd: c())
        # Hover effect
        def _enter(e):
            card_outer._card.configure(bg=T["card"])
            for ch in self._all_widgets(card_outer._inner):
                try:
                    ch.configure(bg=T["card"])
                except Exception:
                    pass
        def _leave(e):
            card_outer._card.configure(bg=T["surface"])
            for ch in self._all_widgets(card_outer._inner):
                try:
                    ch.configure(bg=T["surface"])
                except Exception:
                    pass
        card_outer._card.bind("<Enter>", _enter)
        card_outer._card.bind("<Leave>", _leave)

    def _all_widgets(self, w):
        yield w
        for child in w.winfo_children():
            yield from self._all_widgets(child)

    # ──────────────────────────── HELPERS UI ─────────────────────────

    def _select_tipo(self, value: str):
        T = self._T
        self._tipo.set(value)
        for val, wdg in self._tipo_cards.items():
            if val == value:
                bg  = T["accent"]
                fg  = T["btn_fg"]
                fg2 = "#93c5fd"
                chk = "✓"
                ic_fg = T["btn_fg"]
                border_c = T["accent"]
            else:
                bg  = T["card"]
                fg  = T["text"]
                fg2 = T["muted"]
                chk = ""
                ic_fg = T["muted"]
                border_c = T["border"]
            for w in wdg["all"]:
                try:
                    w.configure(bg=bg)
                except Exception:
                    pass
            wdg["outer"].configure(bg=border_c)
            wdg["main"].configure(fg=fg)
            wdg["sub"].configure(fg=fg2)
            wdg["check"].configure(text=chk)
            wdg["icon"].configure(fg=ic_fg)

        self.cli_label.config(
            text="NÚMERO DE CLIENTE" if value == "tarjeta" else "NÚMERO DE EMPRESA")

    def _toggle_theme(self):
        self._tags = self.tag_input.get_tags() if hasattr(self, "tag_input") else self._tags
        self._theme = "light" if self._theme == "dark" else "dark"
        self._T = THEMES[self._theme]
        self.configure(bg=self._T["bg"])
        for w in self.winfo_children():
            w.destroy()
        self._build_ui()

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Texto", "*.txt"), ("Todos", "*.*")],
            initialfile="layout_omonel.txt",
        )
        if path:
            self._out_path.set(path)

    def _set_btn_state(self, state: str):
        T = self._T
        cfg = {
            "normal":  (T["accent"],   T["btn_fg"], "GENERAR LAYOUT OMONEL  →", "normal"),
            "loading": (T["border"],   T["muted"],  "⏳  Procesando…",           "disabled"),
            "success": (T["success"],  T["btn_fg"], "✓  GENERADO EXITOSAMENTE",  "disabled"),
            "error":   (T["error"],    T["btn_fg"], "✗  ERROR — reintentar",     "normal"),
        }[state]
        self.btn_run.config(bg=cfg[0], fg=cfg[1], text=cfg[2], state=cfg[3],
                            activebackground=cfg[0], activeforeground=cfg[1])

    def _update_summary(self, df, unmatched: int,
                        txt_path: str, xlsx_path: str):
        total = df["importe"].sum()
        self._metric_vars["registros"].set(f"{len(df):,}")
        self._metric_vars["importe"].set(f"${total:,.2f}")
        self._metric_vars["sin_match"].set(f"{unmatched:,}")
        self._update_file_btn_card(self._btn_txt_card,
                                   lambda p=txt_path: open_file(p))
        self._update_file_btn_card(self._btn_xlsx_card,
                                   lambda p=xlsx_path: open_file(p))

    def _show_help(self):
        T   = self._T
        win = tk.Toplevel(self)
        win.title("Manual de Ayuda")
        win.configure(bg=T["bg"])
        win.geometry("560x480")
        win.resizable(False, False)

        tk.Label(win, text="AYUDA  —  DISPERSIÓN OMONEL",
                 bg=T["bg"], fg=T["accent"],
                 font=(_FF, 11, "bold")).pack(padx=20, pady=(16, 8))

        txt = tk.Text(win, bg=T["surface"], fg=T["text"],
                      font=FONT_MONO, relief="flat", bd=0, wrap="word",
                      state="normal",
                      highlightthickness=1, highlightbackground=T["border"])
        txt.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        txt.insert("end", HELP_TEXT)
        txt.config(state="disabled")

        tk.Button(win, text="Cerrar", command=win.destroy,
                  bg=T["accent"], fg=T["btn_fg"],
                  activebackground=T["accent"], activeforeground=T["btn_fg"],
                  font=FONT_SMALL, relief="flat", bd=0,
                  highlightthickness=0, takefocus=0, cursor="hand2",
                  ).pack(pady=(0, 16), ipadx=20, ipady=6)

    # ──────────────────────────── RUN / THREADING ────────────────────

    def _run(self):
        if self._working:
            return

        p_people  = self._people_path.get().strip()
        p_vales   = self._vales_path.get().strip()
        conceptos = self.tag_input.get_tags()
        out_path  = self._out_path.get().strip()
        cliente   = self._cliente.get().strip()
        hdr_str_p = self._header_row.get().strip()
        hdr_str_v = self._header_row_vales.get().strip()

        errors = []
        if not p_people:  errors.append("Selecciona el archivo People.")
        if not p_vales:   errors.append("Selecciona el archivo Cuenta Vales.")
        if not out_path:  errors.append("Define la ruta del archivo de salida.")
        if not cliente:   errors.append("Ingresa el Número de cliente / empresa.")
        elif not cliente.isdigit():
            errors.append("El Número de cliente debe ser solo dígitos.")
        if not hdr_str_p.isdigit() or int(hdr_str_p) < 1:
            errors.append("Fila de encabezados (People) debe ser ≥ 1.")
        if not hdr_str_v.isdigit() or int(hdr_str_v) < 1:
            errors.append("Fila de encabezados (Cuenta Vales) debe ser ≥ 1.")
        if errors:
            messagebox.showerror("Validación", "\n".join(errors))
            return

        self._working = True
        self._set_btn_state("loading")

        log = self.log_panel.log
        log("─" * 48, "")
        log("Iniciando proceso de dispersión…", "info")
        if conceptos:
            log(f"Conceptos a filtrar: {', '.join(conceptos)}", "info")

        params = {
            "p_people":          p_people,
            "p_vales":           p_vales,
            "conceptos":         conceptos,
            "out_path":          out_path,
            "cliente":           cliente,
            "header_row_people": int(hdr_str_p),
            "header_row_vales":  int(hdr_str_v),
            "tipo":              self._tipo.get(),
        }
        threading.Thread(target=self._worker, args=(params,), daemon=True).start()
        self.after(50, self._poll_queue)

    def _worker(self, params):
        q = self._queue

        def log(msg, kind=""):
            q.put({"type": "log", "msg": msg, "kind": kind})

        def progress(step):
            q.put({"type": "progress", "step": step})

        result = process_files(
            params["p_people"], params["p_vales"],
            params["conceptos"], log, progress,
            params["header_row_people"],
            params["header_row_vales"],
        )

        if result is None:
            q.put({"type": "done", "success": False})
            return

        df        = result["df"]
        unmatched = result["unmatched"]
        txt_path  = params["out_path"]
        xlsx_path = os.path.splitext(txt_path)[0] + ".xlsx"
        errors    = []

        progress(2)

        try:
            if params["tipo"] == "tarjeta":
                generate_omonel_txt(df, txt_path, params["cliente"])
                log(f"TXT (Por Tarjeta) → {txt_path}", "ok")
            else:
                generate_omonel_txt_empleado(df, txt_path, params["cliente"])
                log(f"TXT (Por Empleado) → {txt_path}", "ok")
        except Exception as e:
            log(f"Error al guardar TXT: {e}", "err")
            errors.append(str(e))

        try:
            generate_omonel_layout(df, xlsx_path)
            log(f"Excel (referencia) → {xlsx_path}", "ok")
        except Exception as e:
            log(f"Error al guardar Excel: {e}", "err")
            errors.append(str(e))

        total = df["importe"].sum()
        log(f"Layout Omonel generado con éxito.", "ok")
        log(f"Resumen de datos disponible para revisión.", "info")
        progress(3)

        q.put({
            "type":      "done",
            "success":   len(errors) == 0,
            "df":        df,
            "unmatched": unmatched,
            "txt_path":  txt_path,
            "xlsx_path": xlsx_path,
            "errors":    errors,
        })

    def _poll_queue(self):
        try:
            while True:
                msg   = self._queue.get_nowait()
                mtype = msg["type"]

                if mtype == "log":
                    self.log_panel.log(msg["msg"], msg["kind"])

                elif mtype == "progress":
                    pass

                elif mtype == "done":
                    self._working = False
                    log = self.log_panel.log
                    log("─" * 48, "")

                    if msg["success"]:
                        self._set_btn_state("success")
                        self.after(3000, lambda: self._set_btn_state("normal"))
                        self._update_summary(
                            msg["df"], msg["unmatched"],
                            msg["txt_path"], msg["xlsx_path"],
                        )
                    else:
                        self._set_btn_state("error")
                        if msg.get("errors"):
                            messagebox.showerror(
                                "Errores al guardar",
                                "\n".join(msg["errors"]),
                            )
                    return

        except queue.Empty:
            pass

        if self._working:
            self.after(50, self._poll_queue)


def main():
    app = OmonelApp()
    app.mainloop()


if __name__ == "__main__":
    main()
